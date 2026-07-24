"""CSV/XLSX parsing and export for database transfer jobs."""

import csv
import io
import json
import uuid
from datetime import date, datetime
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.database import Database
from app.models.field import Entity, Field, FieldType
from app.services.engine import CellValidationError, next_entity_seq, validate_required_fields

MAX_IMPORT_ROWS = 100_000
FORMULA_PREFIXES = ("=", "+", "-", "@")


def read_tabular(data: bytes, file_format: str) -> tuple[list[str], list[list[Any]]]:
    records: list[list[Any]]
    if file_format == "csv":
        reader = csv.reader(io.StringIO(data.decode("utf-8-sig")))
        records = list(reader)
    else:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = cast(Worksheet, workbook.active)
        records = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    if not records:
        return [], []
    headers = [str(value or "").strip() for value in records[0]]
    rows = records[1 : MAX_IMPORT_ROWS + 1]
    return headers, rows


def _infer_type(values: list[Any]) -> FieldType:
    present = [value for value in values if value not in (None, "")]
    if present and all(
        isinstance(value, int | float) and not isinstance(value, bool) for value in present
    ):
        return FieldType.number
    return FieldType.text


def _export_value(value: Any) -> str | int | float | bool | date | datetime | None:
    """Convert JSONB cell values into safe, portable spreadsheet values."""
    if value is None or isinstance(value, int | float | bool | date | datetime):
        return value
    if isinstance(value, str):
        # Prevent imported user content from becoming an executable spreadsheet formula.
        return f"'{value}" if value.startswith(FORMULA_PREFIXES) else value
    if isinstance(value, dict):
        start = value.get("start")
        end = value.get("end")
        if start is not None and set(value).issubset({"start", "end"}):
            return f"{start} → {end}" if end else str(start)
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    if isinstance(value, list):
        if all(item is None or isinstance(item, str | int | float | bool) for item in value):
            return ", ".join("" if item is None else str(item) for item in value)
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return str(value)


async def import_entities(
    db: AsyncSession,
    *,
    database: Database,
    headers: list[str],
    records: list[list[Any]],
    mapping: dict[str, str],
    field_types: dict[str, str],
    create_missing_fields: bool,
    data_source_id: uuid.UUID,
    name_column: str,
    include_rows: list[int] | None = None,
    incoming_duplicate_policy: str = "suffix",
    existing_name_policy: str = "suffix",
) -> dict[str, int]:
    result = await db.execute(
        select(Field).where(Field.database_id == database.id).order_by(Field.order)
    )
    fields = list(result.scalars())
    by_name = {field.name.casefold(): field for field in fields}
    by_id = {str(field.id): field for field in fields}
    if name_column not in headers:
        raise ValueError("A Name column must be selected before importing")
    name_index = headers.index(name_column)
    system_uid = next((f for f in fields if (f.options or {}).get("system_key") == "uid"), None)
    system_name = next((f for f in fields if (f.options or {}).get("system_key") == "name"), None)
    if system_name is None:
        raise ValueError("Database is missing its built-in Name field")

    original_field_count = len(fields)
    column_fields: list[Field | None] = []
    for index, header in enumerate(headers):
        mapped = by_id.get(mapping.get(header, ""))
        field = mapped or by_name.get(header.casefold())
        if field is None and create_missing_fields and header:
            values = [row[index] if index < len(row) else None for row in records]
            field = Field(
                database_id=database.id,
                name=header,
                type=FieldType(field_types.get(header, _infer_type(values))),
                options={},
                order=len(fields),
            )
            db.add(field)
            await db.flush()
            fields.append(field)
            by_name[header.casefold()] = field
        # UID/Name are system identity, not ordinary import target columns.
        column_fields.append(
            None if field is system_uid or field is system_name else field
        )

    auto_types = {
        FieldType.unique_id,
        FieldType.rollup,
        FieldType.formula,
        FieldType.created_time,
        FieldType.created_by,
        FieldType.last_edited_time,
        FieldType.last_edited_by,
    }
    required_fields = [
        field
        for field in fields
        if field.type not in auto_types
        and (field.options or {}).get("required") is True
        and field is not system_name
    ]
    mapped_field_ids = {field.id for field in column_fields if field is not None}
    missing_mappings = [field.name for field in required_fields if field.id not in mapped_field_ids]
    if missing_mappings:
        raise ValueError(
            f"Required field mapping missing: {', '.join(missing_mappings)}"
        )

    selected = set(include_rows) if include_rows is not None else None
    indexed_records = [
        (index, record)
        for index, record in enumerate(records)
        if selected is None or index in selected
    ]
    entities_result = await db.execute(select(Entity).where(Entity.database_id == database.id))
    existing_by_name: dict[str, Entity] = {
        entity.name.casefold(): entity for entity in entities_result.scalars()
    }
    original_existing_by_name = dict(existing_by_name)
    taken_names = set(existing_by_name)
    seen_incoming: set[str] = set()

    def reserve_name(raw: Any) -> str:
        base_name = str(raw or "").strip()
        if not base_name:
            raise ValueError("Every imported entity needs a Name")
        if len(base_name) > 200:
            base_name = base_name[:200]
        candidate = base_name
        suffix = 2
        while candidate.casefold() in taken_names:
            candidate = f"{base_name[:190]} {suffix}"
            suffix += 1
        taken_names.add(candidate.casefold())
        return candidate

    base = await next_entity_seq(db, database.id) - 1
    imported = updated = skipped = suffixed = 0
    for _row_index, record in indexed_records:
        source_name = record[name_index] if name_index < len(record) else None
        raw_name = str(source_name or "").strip()
        if not raw_name:
            skipped += 1
            continue
        key = raw_name.casefold()
        duplicate_in_file = key in seen_incoming
        seen_incoming.add(key)
        if duplicate_in_file and incoming_duplicate_policy == "skip":
            skipped += 1
            continue
        data: dict[str, Any] = {}
        for index, field in enumerate(column_fields):
            if field is None or index >= len(record):
                continue
            value = record[index]
            if value in (None, ""):
                continue
            data[str(field.id)] = value
        existing = original_existing_by_name.get(key)
        if existing is not None and existing_name_policy == "update":
            next_data = {**existing.data, **data, str(system_name.id): existing.name}
            try:
                validate_required_fields(fields, next_data)
            except CellValidationError as exc:
                raise ValueError(f"Row {_row_index + 2}: {exc}") from exc
            existing.data = next_data
            updated += 1
            continue
        name = reserve_name(raw_name)
        if name.casefold() != key:
            suffixed += 1
        seq = base + imported + 1
        uid = str(seq)
        data[str(system_name.id)] = name
        if system_uid is not None:
            data[str(system_uid.id)] = uid
        try:
            validate_required_fields(fields, data)
        except CellValidationError as exc:
            raise ValueError(f"Row {_row_index + 2}: {exc}") from exc
        entity = Entity(
            database_id=database.id,
            data_source_id=data_source_id,
            data=data,
            uid=uid,
            name=name,
            seq=seq,
            order=seq,
        )
        db.add(entity)
        existing_by_name[name.casefold()] = entity
        imported += 1
        if (imported + updated) % 1000 == 0:
            await db.flush()
    await db.commit()
    return {
        "entities_imported": imported,
        "entities_updated": updated,
        "entities_skipped": skipped,
        "entities_suffixed": suffixed,
        "fields_created": len(fields) - original_field_count,
    }


def export_entities(
    fields: list[Field], entities: list[Entity], file_format: str
) -> tuple[bytes, str]:
    headers = [field.name for field in fields]
    values = [
        [_export_value(entity.data.get(str(field.id))) for field in fields] for entity in entities
    ]
    if file_format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(values)
        return output.getvalue().encode("utf-8-sig"), "text/csv"

    output_bytes = io.BytesIO()
    workbook = Workbook(write_only=False)
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Data"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for row_values in values:
        sheet.append(row_values)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        width = min(50, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(output_bytes)
    return (
        output_bytes.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
